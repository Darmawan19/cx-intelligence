"""
report_generator.py — Step 6 of CX Intelligence Pipeline

Produces a professional Excel workbook with:
  Sheet 1: Executive Summary (text + key stats)
  Sheet 2: Priority Matrix (ranked issue areas)
  Sheet 3: Issue Type Breakdown
  Sheet 4: Deep Dive 1 — Top Issue
  Sheet 5: Deep Dive 2 — 2nd Issue  
  Sheet 6: Deep Dive 3 — 3rd Issue
  Sheet 7: Raw Classified Data
  Sheet 8: Methodology

This is the "consumer listening report" output expected by the role.
"""

import os
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint
from src.analyzer import FEATURE_AREA_LABELS, ISSUE_TYPE_LABELS

# ── Color Palette ──────────────────────────────────────────────────────────────
SHOPEE_ORANGE = "EE4D2D"
SHOPEE_ORANGE_LIGHT = "FDECEA"
DARK_HEADER = "1A1A2E"
SECTION_BG = "F0F2F5"
WHITE = "FFFFFF"
TEXT_DARK = "1A1A2E"
TEXT_MEDIUM = "4A4A5A"
SEVERITY_COLORS = {
    5: "C0392B",  # Critical — dark red
    4: "E74C3C",  # High — red
    3: "F39C12",  # Medium — orange
    2: "F1C40F",  # Low-med — yellow
    1: "2ECC71",  # Low — green
}


def _make_border(style="thin"):
    s = Side(style=style, color="D0D0D0")
    return Border(left=s, right=s, top=s, bottom=s)


def _header_style(ws, cell, text, size=12, bold=True, bg=DARK_HEADER, fg=WHITE, center=True):
    c = ws[cell]
    c.value = text
    c.font = Font(name="Calibri", bold=bold, size=size, color=fg)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center" if center else "left", vertical="center", wrap_text=True)
    return c


def _subheader(ws, cell, text, size=10, color=TEXT_DARK):
    c = ws[cell]
    c.value = text
    c.font = Font(name="Calibri", bold=True, size=size, color=color)
    c.fill = PatternFill("solid", fgColor=SECTION_BG)
    c.alignment = Alignment(horizontal="left", vertical="center")
    return c


def _data_cell(ws, cell, value, center=False, number_format=None, color=TEXT_DARK, bold=False):
    c = ws[cell]
    c.value = value
    c.font = Font(name="Calibri", size=10, color=color, bold=bold)
    c.alignment = Alignment(horizontal="center" if center else "left", vertical="center", wrap_text=True)
    if number_format:
        c.number_format = number_format
    return c


def _sheet1_executive(wb: Workbook, exec_summary: str, summary: dict, area_stats: pd.DataFrame):
    ws = wb.active
    ws.title = "Executive Summary"
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 3

    ws.row_dimensions[1].height = 8

    # Title
    ws.merge_cells("B2:E2")
    _header_style(ws, "B2", "SHOPEE INDONESIA — CONSUMER LISTENING REPORT", size=14)

    ws.merge_cells("B3:E3")
    _header_style(ws, "B3",
                  f"CX Intelligence Analysis | {summary['date_range_start']} to {summary['date_range_end']} | Generated: {summary['generated_at']}",
                  size=9, bg=SHOPEE_ORANGE, fg=WHITE)

    # Stats row
    ws.row_dimensions[5].height = 50
    stats = [
        ("Reviews Analyzed", f"{summary['total_reviews']:,}", "B5", "B6"),
        ("Negative Reviews", f"{summary['pct_negative']}%", "C5", "C6"),
        ("Avg Severity", f"{summary['avg_severity_overall']}/5.0", "D5", "D6"),
        ("Issue Areas Found", str(len(area_stats[area_stats['feature_area'] != 'other'])), "E5", "E6"),
    ]
    for label, val, cell_v, cell_l in stats:
        ws.merge_cells(f"{cell_v}:{cell_v}")
        c = ws[cell_v]
        c.value = val
        c.font = Font(name="Calibri", bold=True, size=20, color=SHOPEE_ORANGE)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.fill = PatternFill("solid", fgColor=SHOPEE_ORANGE_LIGHT)

        cl = ws[cell_l]
        cl.value = label
        cl.font = Font(name="Calibri", size=9, color=TEXT_MEDIUM)
        cl.alignment = Alignment(horizontal="center", vertical="top")
        cl.fill = PatternFill("solid", fgColor=SHOPEE_ORANGE_LIGHT)

    # Executive Summary text
    ws.row_dimensions[8].height = 15
    ws.merge_cells("B8:E8")
    _subheader(ws, "B8", "  EXECUTIVE SUMMARY", size=10)

    ws.merge_cells("B9:E18")
    c = ws["B9"]
    c.value = exec_summary
    c.font = Font(name="Calibri", size=10, color=TEXT_DARK)
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.row_dimensions[9].height = 120

    # Top 3 Issues box
    ws.row_dimensions[20].height = 15
    ws.merge_cells("B20:E20")
    _subheader(ws, "B20", "  TOP 3 PRIORITY ISSUES", size=10)

    headers = ["Rank", "Issue Area", "Volume", "Priority Score"]
    for i, h in enumerate(headers, 2):
        col = get_column_letter(i)
        _header_style(ws, f"{col}21", h, size=9, bg="3D3D5C", fg=WHITE)

    top3 = area_stats[area_stats["feature_area"] != "other"].head(3)
    for row_idx, (_, row) in enumerate(top3.iterrows(), 22):
        rank_colors = ["C0392B", "E67E22", "27AE60"]
        bg = rank_colors[row_idx - 22]
        _data_cell(ws, f"B{row_idx}", f"#{row_idx-21}", center=True, bold=True, color=bg)
        _data_cell(ws, f"C{row_idx}", row["label"])
        _data_cell(ws, f"D{row_idx}", int(row["volume"]), center=True)
        _data_cell(ws, f"E{row_idx}", row["priority_score"], center=True, number_format="0.0")
        ws.row_dimensions[row_idx].height = 20


def _sheet2_priority_matrix(wb: Workbook, area_stats: pd.DataFrame):
    ws = wb.create_sheet("Priority Matrix")
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 16
    ws.column_dimensions["G"].width = 16
    ws.column_dimensions["H"].width = 14

    ws.merge_cells("B1:H1")
    _header_style(ws, "B1", "ISSUE PRIORITY MATRIX — Ranked by Priority Score", size=12)

    headers = ["Rank", "Issue Area", "Volume", "% of Total", "Avg Severity", "% Widespread", "Priority Score", "Recommended Action"]
    for i, h in enumerate(headers, 2):
        col = get_column_letter(i)
        _header_style(ws, f"{col}2", h, size=9, bg="3D3D5C")

    # Methodology note
    ws["B3"].value = (
        "Priority Score = Volume × Avg Severity × (1 + Widespread Bonus)  |  "
        "Widespread Bonus: +50% max if many users report systemic impact"
    )
    ws["B3"].font = Font(name="Calibri", size=8, italic=True, color=TEXT_MEDIUM)
    ws.merge_cells("B3:H3")

    for row_idx, (_, row) in enumerate(area_stats.iterrows(), 4):
        bg = "FFF5F5" if row["rank"] <= 3 else "FAFAFA"
        rank_text = f"#{int(row['rank'])}"

        _data_cell(ws, f"B{row_idx}", rank_text, center=True, bold=(row["rank"] <= 3))
        _data_cell(ws, f"C{row_idx}", row["label"])
        _data_cell(ws, f"D{row_idx}", int(row["volume"]), center=True)
        _data_cell(ws, f"E{row_idx}", f"{row['pct_of_total']}%", center=True)

        # Severity with color
        sev_cell = ws[f"F{row_idx}"]
        sev_cell.value = f"{row['avg_severity']:.2f} / 5.0"
        sev_int = round(row["avg_severity"])
        sev_cell.fill = PatternFill("solid", fgColor=SEVERITY_COLORS.get(sev_int, "AAAAAA") + "22")
        sev_cell.font = Font(name="Calibri", size=10, color=SEVERITY_COLORS.get(sev_int, "333333"), bold=True)
        sev_cell.alignment = Alignment(horizontal="center", vertical="center")

        _data_cell(ws, f"G{row_idx}", f"{row['pct_widespread']:.1f}%", center=True)
        _data_cell(ws, f"H{row_idx}", round(row["priority_score"], 1), center=True,
                   bold=(row["rank"] <= 3), color=SHOPEE_ORANGE if row["rank"] <= 3 else TEXT_DARK)

        action = "→ IMMEDIATE ACTION" if row["rank"] <= 3 else ("→ MONITOR" if row["rank"] <= 6 else "→ BACKLOG")
        _data_cell(ws, f"I{row_idx}", action, color=SHOPEE_ORANGE if row["rank"] <= 3 else TEXT_MEDIUM)

        ws.row_dimensions[row_idx].height = 20


def _sheet_deep_dive(wb: Workbook, area: str, brief: str, sheet_num: int):
    label = FEATURE_AREA_LABELS.get(area, area)
    ws = wb.create_sheet(f"Deep Dive {sheet_num} — {label[:15]}")
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 100

    ws.merge_cells("B1:B1")
    _header_style(ws, "B1", f"DEEP DIVE #{sheet_num}: {label.upper()}", size=12)

    ws["B2"].value = brief
    ws["B2"].font = Font(name="Calibri", size=10, color=TEXT_DARK)
    ws["B2"].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.row_dimensions[2].height = 500


def _sheet_raw_data(wb: Workbook, df: pd.DataFrame):
    ws = wb.create_sheet("Raw Classified Data")

    cols = [
        "review_index", "review_id", "date", "rating", "review_text",
        "feature_area", "issue_type", "severity", "sentiment",
        "key_phrase", "user_impact", "thumbs_up"
    ]
    cols = [c for c in cols if c in df.columns]

    headers = [c.replace("_", " ").title() for c in cols]
    for i, h in enumerate(headers, 1):
        col = get_column_letter(i)
        _header_style(ws, f"{col}1", h, size=9, bg=DARK_HEADER)

    for row_idx, (_, row) in enumerate(df[cols].iterrows(), 2):
        for col_idx, col in enumerate(cols, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            val = row[col]
            if hasattr(val, 'strftime'):
                val = val.strftime("%Y-%m-%d")
            cell.value = str(val) if val is not None else ""
            cell.font = Font(name="Calibri", size=9)
            cell.alignment = Alignment(wrap_text=(col == "review_text"), vertical="top")

        ws.row_dimensions[row_idx].height = 15 if "review_text" not in cols else 30


def _sheet_methodology(wb: Workbook, summary: dict, quality_report: dict):
    ws = wb.create_sheet("Methodology")
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 65

    ws.merge_cells("B1:C1")
    _header_style(ws, "B1", "RESEARCH METHODOLOGY & TRANSPARENCY NOTE", size=12)

    sections = [
        ("DATA SOURCE", [
            ("Platform", "Google Play Store — Shopee Indonesia (com.shopee.id)"),
            ("Period", f"{summary['date_range_start']} to {summary['date_range_end']}"),
            ("Total Reviews", f"{summary['total_reviews']:,} after deduplication and cleaning"),
            ("Sampling", "Stratified by rating: 30% (1★), 25% (2★), 20% (3★), 15% (4★), 10% (5★)"),
            ("Filter", "Reviews ≥15 characters; last 6 months; deduplicated by review ID"),
        ]),
        ("CLASSIFICATION", [
            ("Method", "AI-augmented batch classification via Claude API (anthropic)"),
            ("Batch Size", "20 reviews per API call"),
            ("Fields", "feature_area, issue_type, severity (1-5), sentiment, key_phrase, user_impact"),
            ("Quality Gate", f"Random sample of {quality_report.get('sample_size', 50)} reviews validated"),
            ("Error Rate", f"{quality_report.get('error_rate', 0):.1%} (threshold: {quality_report.get('threshold', 0.15):.0%})"),
            ("Gate Status", "PASSED" if quality_report.get("passed") else "FAILED — review with caution"),
        ]),
        ("PRIORITIZATION", [
            ("Framework", "Modified RICE: Priority Score = Volume × Avg Severity × (1 + Widespread Bonus)"),
            ("Widespread Bonus", "Up to +50% for issues flagged as 'likely_widespread' by classifier"),
            ("Excluded", "'Other/Unclassified' excluded from top-3 ranking"),
            ("Effort", "Not included in score — researcher role flags issues, not fix cost"),
        ]),
        ("LIMITATIONS", [
            ("Single Source", "Google Play only; iOS App Store users may have different pain points"),
            ("Language", "Bahasa Indonesia primary; English-language reviews may be underrepresented"),
            ("Recency Bias", "Newest reviews weighted higher; older patterns may be missed"),
            ("AI Accuracy", f"Classification error rate ~{quality_report.get('error_rate', 0):.1%} on validation sample"),
            ("Selection Bias", "Highly satisfied users less likely to leave reviews"),
        ]),
    ]

    row_idx = 2
    for section_title, items in sections:
        ws.merge_cells(f"B{row_idx}:C{row_idx}")
        _subheader(ws, f"B{row_idx}", f"  {section_title}", size=10)
        ws.row_dimensions[row_idx].height = 20
        row_idx += 1

        for label, value in items:
            _data_cell(ws, f"B{row_idx}", label, bold=True, color=TEXT_MEDIUM)
            _data_cell(ws, f"C{row_idx}", value)
            ws.row_dimensions[row_idx].height = 18
            row_idx += 1

        row_idx += 1


def generate_excel_report(
    results: dict,
    deep_dives: dict,
    exec_summary: str,
    quality_report: dict,
    output_path: str = "outputs/shopee_cx_report.xlsx",
) -> str:
    print(f"\n{'='*60}")
    print(f"[STEP 6 / REPORT GENERATOR] Building Excel report")
    print(f"{'='*60}")

    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    wb = Workbook()
    summary = results["summary"]
    area_stats = results["area_stats"]
    df = results["df"]

    print("  Building Sheet 1: Executive Summary...")
    _sheet1_executive(wb, exec_summary, summary, area_stats)

    print("  Building Sheet 2: Priority Matrix...")
    _sheet2_priority_matrix(wb, area_stats)

    print("  Building Deep Dive sheets...")
    for i, (area, brief) in enumerate(deep_dives.items(), 1):
        _sheet_deep_dive(wb, area, brief, i)
        print(f"    Sheet {i+2}: Deep Dive — {FEATURE_AREA_LABELS.get(area, area)}")

    print("  Building Raw Data sheet...")
    _sheet_raw_data(wb, df)

    print("  Building Methodology sheet...")
    _sheet_methodology(wb, summary, quality_report)

    wb.save(output_path)
    print(f"\n[REPORT] ✓ Saved → {output_path}")
    return output_path
